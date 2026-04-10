"""Geração de relatórios consolidados HTML + Excel."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import logging

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import BASE_DIR, settings
from app.models import OcOrfa, ResultadoValidacao, StatusValidacao


_jinja = Environment(
    loader=FileSystemLoader(BASE_DIR / "templates"),
    autoescape=select_autoescape(["html"]),
)


def _acao_sugerida(r: ResultadoValidacao) -> str:
    """Texto curto do que o sistema FARIA em modo automático."""
    s = r.status
    if s == StatusValidacao.APROVADA:
        return "Aprovar"
    if s == StatusValidacao.DIVERGENCIA:
        return "Bloquear / Revisar"
    if s == StatusValidacao.BLOQUEADA:
        return "Bloquear"
    if s == StatusValidacao.AGUARDANDO_ML:
        return "Aguardar análise ML"
    if s == StatusValidacao.JA_PROCESSADA:
        return "Já processada"
    return "—"


def _link_card(card_id: str | None) -> str:
    if not card_id:
        return ""
    return f"https://app.pipefy.com/pipes/{settings.pipe_id}#cards/{card_id}"


def gerar_html(
    data_d1: date,
    resultados: list[ResultadoValidacao],
    *,
    dry_run: bool,
    ocs_orfas: list[OcOrfa] | None = None,
    historico_status: dict | None = None,
) -> Path:
    template = _jinja.get_template("relatorio.html.j2")
    total = len(resultados)
    aprovadas = sum(1 for r in resultados if r.status == StatusValidacao.APROVADA)
    divergentes = sum(1 for r in resultados if r.status == StatusValidacao.DIVERGENCIA)
    bloqueadas = sum(1 for r in resultados if r.status == StatusValidacao.BLOQUEADA)
    aguardando_ml = sum(1 for r in resultados if r.status == StatusValidacao.AGUARDANDO_ML)
    ja_processadas = sum(1 for r in resultados if r.status == StatusValidacao.JA_PROCESSADA)

    html = template.render(
        data_d1=data_d1.isoformat(),
        gerado_em=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        executado_por=settings.validador_identificador,
        dry_run=dry_run,
        modo_operacao=settings.modo_operacao,
        total=total,
        aprovadas=aprovadas,
        divergentes=divergentes,
        bloqueadas=bloqueadas,
        aguardando_ml=aguardando_ml,
        ja_processadas=ja_processadas,
        resultados=resultados,
        ocs_orfas=ocs_orfas or [],
        pipe_id=settings.pipe_id,
        acao_sugerida=_acao_sugerida,
        link_card=_link_card,
        cilia_mode=settings.cilia_mode,
        cilia_base_url=settings.cilia_base_url,
        historico_status=historico_status,
    )

    out = settings.relatorios_full_dir / f"{data_d1.isoformat()}_validacao.html"
    out.write_text(html, encoding="utf-8")
    return out


def gerar_excel(
    data_d1: date,
    resultados: list[ResultadoValidacao],
    *,
    ocs_orfas: list[OcOrfa] | None = None,
    historico_status: dict | None = None,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Validação {data_d1.isoformat()}"

    # ----- Banner de histórico incompleto (linha 1, mesclada) -----
    # Aparece como primeira linha (antes dos headers) quando o histórico
    # está incompleto, sinalizando para o analista que os alertas de
    # duplicidade podem estar parciais.
    banner_row = 0
    if historico_status and not historico_status.get("completo", True):
        cobertos = historico_status.get("dias_cobertos", 0)
        necessarios = historico_status.get("dias_necessarios", 0)
        banner_text = (
            f"⚠ HISTÓRICO DE DUPLICIDADES INCOMPLETO — "
            f"{cobertos}/{necessarios} dias populados. "
            f"Alertas de peças duplicadas podem estar incompletos. "
            f"O sistema continuará baixando o histórico nas próximas execuções."
        )
        ws.append([banner_text])
        banner_row = 1
        banner_cell = ws.cell(row=1, column=1)
        banner_cell.font = Font(bold=True, color="991B1B", size=12)
        banner_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        banner_cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[1].height = 32

    headers = [
        "Nº OC",
        "Placa (com hífen)",
        "Placa (sem hífen)",
        "Fornecedor",
        "Comprador",
        "Forma Pagamento",
        "Valor Card (Pipefy)",
        "Valor OC (Club)",
        "Valor PDF (Pipefy)",
        "Valor Cilia",
        "Qtd Cotações",
        "Qtd Produtos",
        "Peça Duplicada",
        "Reincidência (7m)",
        "Detalhe Reincidência",
        "Cancelamento",
        "Link Cancelamento",
        "Cília (verificar)",
        "Status",
        "Ação Sugerida",
        "Card Pipefy",
        "Motivo",
        "Fase Pipefy Atual",
        "Fase Pipefy Destino",
    ]
    ws.append(headers)
    header_row_idx = banner_row + 1  # linha onde os headers foram inseridos
    data_start_row = header_row_idx + 1

    # Mescla a célula do banner acima de todos os headers
    if banner_row:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(headers),
        )

    header_fill = PatternFill("solid", fgColor="1A2332")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")

    status_fill = {
        "aprovada": PatternFill("solid", fgColor="D1FAE5"),
        "divergencia": PatternFill("solid", fgColor="FED7AA"),
        "bloqueada": PatternFill("solid", fgColor="FECACA"),
        "aguardando_ml": PatternFill("solid", fgColor="FEF3C7"),
        "ja_processada": PatternFill("solid", fgColor="E5E7EB"),
    }
    status_label = {
        "aprovada": "Aprovada",
        "divergencia": "Divergência",
        "bloqueada": "Bloqueada",
        "aguardando_ml": "Aguardando ML",
        "ja_processada": "Já processada",
    }

    reinc_label_map = {
        "—": "—",
        "sim_devolucao": "Devolução aberta",
        "sim_com_devolucao_peca": "Devolução da peça ✅",
        "sim_devolucao_outra_peca": "Devolução de OUTRA peça ⚠️",
        "sim_sem_devolucao": "SEM devolução ❌",
        "sim_sem_devolucao_mesmo_forn": "SEM devolução (mesmo forn.) ❌",
        "sim_mesmo_forn": "Mesmo fornecedor",
        "sim_outro_forn": "Outro fornecedor",
    }
    cancel_label_map = {
        "—": "—",
        "info_incorretas": "Em revisão",
        "cancelado": "Cancelado",
        "ambos": "Cancelado (rev.)",
    }
    pipe_id_principal = settings.pipe_id

    for row_idx, r in enumerate(resultados, start=data_start_row):
        forn = r.oc.fornecedor.for_nome if r.oc.fornecedor else ""
        motivo = r.motivo_resumido
        if r.status == StatusValidacao.AGUARDANDO_ML:
            motivo = "Fornecedor Mercado Livre — validação manual do analista"
        elif r.status == StatusValidacao.JA_PROCESSADA:
            motivo = f"Card já estava em fase '{r.fase_pipefy_atual or '?'}'"
        link_pipefy = _link_card(r.card_pipefy_id)

        # Detalhe da reincidência (todas as peças cross-time)
        detalhe_reinc = ""
        if r.divergencias_cross:
            partes = []
            for div in r.divergencias_cross:
                d = div.dados or {}
                status_dev = "✅ devolução" if d.get("tem_devolucao_peca") \
                    else "⚠️ dev. outra peça" if d.get("tem_devolucao_outra_peca") \
                    else "❌ SEM devolução"
                partes.append(
                    f"{d.get('descricao_peca','—')} "
                    f"({d.get('data_anterior','—')} OC {d.get('oc_anterior','—')} "
                    f"forn: {d.get('fornecedor_anterior_nome','—')}) "
                    f"[{status_dev}]"
                )
            detalhe_reinc = " | ".join(partes)

        link_cancel = (
            f"https://app.pipefy.com/pipes/{pipe_id_principal}#cards/{r.cancelamento_card_id}"
            if r.cancelamento_card_id else ""
        )
        # Modo deeplink: link de login do Cilia para validação manual
        link_cilia = (
            f"{settings.cilia_base_url}/users/sign_in"
            if settings.cilia_mode == "deeplink"
            else ""
        )

        ws.append([
            r.oc.id_pedido,
            r.oc.identificador or "",
            r.oc.placa_normalizada,
            forn,
            r.oc.comprador_nome or (f"#{r.oc.created_by}" if r.oc.created_by else ""),
            r.oc.forma or "",
            float(r.valor_card) if r.valor_card else None,
            float(r.valor_club) if r.valor_club else None,
            float(r.valor_pdf) if r.valor_pdf else None,
            float(r.valor_cilia) if r.valor_cilia else None,
            r.qtd_cotacoes,
            r.qtd_produtos,
            r.peca_duplicada,
            reinc_label_map.get(r.reincidencia, r.reincidencia),
            detalhe_reinc,
            cancel_label_map.get(r.cancelamento, r.cancelamento),
            link_cancel,
            link_cilia,
            status_label.get(r.status.value, r.status.value),
            _acao_sugerida(r),
            link_pipefy,
            motivo,
            r.fase_pipefy_atual or "",
            r.fase_destino.value if r.fase_destino else "",
        ])
        fill = status_fill.get(r.status.value)
        if fill:
            # Status agora está na coluna 19 (eram 14, depois 18)
            ws.cell(row=row_idx, column=19).fill = fill
        # Hyperlink na coluna "Card Pipefy" (agora 21, eram 16, depois 20)
        if link_pipefy:
            cell = ws.cell(row=row_idx, column=21)
            cell.hyperlink = link_pipefy
            cell.value = "Abrir"
            cell.font = Font(color="0563C1", underline="single")
        # Hyperlink na coluna "Link Cancelamento" (17)
        if link_cancel:
            cell = ws.cell(row=row_idx, column=17)
            cell.hyperlink = link_cancel
            cell.value = "Abrir →"
            cell.font = Font(color="0563C1", underline="single")
        # Hyperlink na coluna "Cília (verificar)" (18)
        if link_cilia:
            cell = ws.cell(row=row_idx, column=18)
            cell.hyperlink = link_cilia
            cell.value = f"🔗 {r.oc.identificador or r.oc.placa_normalizada}"
            cell.font = Font(color="0563C1", underline="single")

    data_end_row = data_start_row + len(resultados)
    # Formato de moeda nas 4 colunas de valor (7=Card, 8=Club, 9=PDF, 10=Cilia)
    for col in (7, 8, 9, 10):
        for row in range(data_start_row, data_end_row):
            ws.cell(row=row, column=col).number_format = "R$ #,##0.00"

    # Larguras automáticas
    for col_idx, header in enumerate(headers, start=1):
        largura = max(
            len(header),
            max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(data_start_row, data_end_row)),
                default=0,
            ),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(largura + 2, 50)

    # Freeze abaixo da linha de headers (se tem banner, congela abaixo da linha 2)
    ws.freeze_panes = f"A{data_start_row}"

    # ----- Aba 2: OCs órfãs (Club sem card no Pipefy) -----
    if ocs_orfas:
        ws2 = wb.create_sheet("Revisão final")
        orfa_headers = [
            "Nº Pedido",
            "Cotação",
            "Placa",
            "Fornecedor",
            "Comprador",
            "Forma Pagamento",
            "Valor Club",
            "Data Pedido",
            "Qtd Itens",
            "Peça Duplicada",
            "Reincidência (7m)",
            "Detalhe Reincidência",
            "Link Verificação",
            "Cancelamento",
            "Link Cancelamento",
            "Cília (verificar)",
        ]
        ws2.append(orfa_headers)
        for col_idx in range(1, len(orfa_headers) + 1):
            c = ws2.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="left", vertical="center")
        # Reusar os mesmos maps definidos na aba 1 (já estão no escopo da função)
        for o in ocs_orfas:
            reinc_label = reinc_label_map.get(o.reincidencia, o.reincidencia)
            cancel_label = cancel_label_map.get(o.cancelamento, o.cancelamento)
            # Detalhe com TODAS as peças reincidentes (não só a última)
            detalhe = ""
            link = ""
            if o.divergencias:
                partes = []
                for div in o.divergencias:
                    d = div.dados or {}
                    status_dev = "✅ devolução" if d.get("tem_devolucao_peca") \
                        else "⚠️ dev. outra peça" if d.get("tem_devolucao_outra_peca") \
                        else "❌ SEM devolução"
                    partes.append(
                        f"{d.get('descricao_peca','—')} "
                        f"({d.get('data_anterior','—')} OC {d.get('oc_anterior','—')} "
                        f"forn: {d.get('fornecedor_anterior_nome','—')}) "
                        f"[{status_dev}]"
                    )
                detalhe = " | ".join(partes)
                d0 = o.divergencias[0].dados or {}
                link = d0.get("link_devolucao") or d0.get("link_oc_anterior") or ""
            link_cancel = (
                f"https://app.pipefy.com/pipes/{pipe_id_principal}#cards/{o.cancelamento_card_id}"
                if o.cancelamento_card_id else ""
            )
            link_cilia_orfa = (
                f"{settings.cilia_base_url}/users/sign_in"
                if settings.cilia_mode == "deeplink"
                else ""
            )
            ws2.append([
                o.id_pedido,
                o.id_cotacao or "",
                o.identificador or "",
                o.fornecedor or "",
                o.comprador or "",
                o.forma_pagamento or "",
                float(o.valor) if o.valor else None,
                o.data_pedido.isoformat() if o.data_pedido else "",
                o.qtd_produtos if o.qtd_produtos is not None else "",
                o.peca_duplicada,
                reinc_label,
                detalhe,
                link,
                cancel_label,
                link_cancel,
                link_cilia_orfa,
            ])
        # Colunas com URL → converter em hyperlink real:
        #   13 = "Link Verificação" (devolução / OC anterior)
        #   15 = "Link Cancelamento"
        #   16 = "Cília (verificar)"
        url_cols = [13, 15, 16]
        for row in range(2, len(ocs_orfas) + 2):
            ws2.cell(row=row, column=7).number_format = "R$ #,##0.00"
            placa_orfa = ws2.cell(row=row, column=3).value or ""
            for col in url_cols:
                cell_link = ws2.cell(row=row, column=col)
                url = cell_link.value
                if url and isinstance(url, str) and url.startswith("http"):
                    cell_link.hyperlink = url
                    if col == 16:
                        cell_link.value = f"🔗 {placa_orfa}"
                    else:
                        cell_link.value = "Abrir →"
                    cell_link.style = "Hyperlink"
        for col_idx, h in enumerate(orfa_headers, start=1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = max(len(h), 14) + 2
        ws2.freeze_panes = "A2"

    out = settings.relatorios_full_dir / f"{data_d1.isoformat()}_validacao.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        # Arquivo aberto no Excel — salva com timestamp e avisa
        ts = datetime.now().strftime("%H%M%S")
        out = settings.relatorios_full_dir / f"{data_d1.isoformat()}_validacao_{ts}.xlsx"
        wb.save(out)
        logging.getLogger(__name__).warning(
            "Arquivo Excel estava aberto — salvo como %s", out.name
        )
    return out
